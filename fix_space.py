import pandas as pd
import numpy as np
import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def key_field_col(sheet):
    key_col=[]
    for header2 in sheet.iter_rows(2,2): # for row in ws.iter_rows(): when includes the sheetsnumber
        for attribute in header2:
            if attribute.fill.start_color.index == 43 or attribute.fill.start_color.index=='FFFFFF99':
                key_col.append(attribute.col_idx)
    return key_col

def check_datatype(sheet):
    Datatype_Int = []
    Datatype_Bol = []
    for header1 in sh.iter_rows(1,1):
        for datatype in header1:
            if datatype.value == 'INTEGER':
                Datatype_Int.append(datatype.col_idx)
            if datatype.value == 'BOOLEAN':
                Datatype_Bol.append(datatype.col_idx)
    return Datatype_Int, Datatype_Bol

def check_only_spaces(cell_value):
    cell_value = str(cell_value)
    if len(cell_value)!= 0: #not empty
        for i in list(cell_value):
            if i != ' ': #not only space
                return False
            else:
                return True #only space
    if len(cell_value)== 0: #empty
        return ('empty')

path1 = '/Users/jingjingshao/OneDrive - *****/Excel_python/'
filename = 'DataDefinition.xlsx'
#sheetname = 'CreditRiskExposure'


if len(sys.argv) > 1:
    coloring = 'yes'
else:
    coloring = 'no'


orginal= os.path.join(path1,filename)
df1 = load_workbook(orginal, data_only = True)
#sh = df1[sheetname]
#for ws in df1.worksheets:

#different Colors for differnt mistakes
color_codes  = ['FC2C03', '03FCF4', '35FC03', 'FCBA03', '008B8B']
Colors = []
for i in color_codes:
    Colors.append(PatternFill(patternType='solid', fgColor=i))

for sh in df1.worksheets:
    mistake_index = [[] for i in range(5)]
    key_col= key_field_col(sh)
    Datatype_Int, Datatype_Bol = check_datatype(sh)
    for row in sh.iter_rows(0,sh.max_row):
        for cell in row:
            if cell.col_idx in key_col:
                if pd.isna(cell.value) == True:#and empty
                    cell.value = ' '
                #if check_only_spaces(cell.value) == 'empty':
                #    cell.value = ' '
                    mistake_index[0].append(cell.coordinate)
                if check_only_spaces(cell.value) == True:#only space
                    if cell.col_idx in Datatype_Int:
                        cell.value = '-1'
                        mistake_index[1].append(cell.coordinate)
                    if cell.col_idx in Datatype_Bol:
                        cell.value = '0'
                        mistake_index[2].append(cell.coordinate)

                if check_only_spaces(cell.value) == False:
                    cell_value = str(cell.value).replace(' ', '')
                    if len(cell_value) != len(str(cell.value)):
                        cell.value = cell_value
                        mistake_index[3].append(cell.coordinate)
            else:
                pass
                if check_only_spaces(cell.value) == True:
                    cell.value = np.nan
                else:
                    cell.value = cell.value

    if coloring == 'yes':
        for i in range(5):
            for index in mistake_index[i]:
                sh[index].fill = Colors[i]

    f = open(os.path.join(path1,"mistake_reports.txt"), "a")
    f.write('{}\n{}\n{}\n{}\n{}\n'.format("Changes detected in sheet: %s\n\n" %sh,
                              "The following changes have been made in the modified file: \n",
                              "Added space in the key fields: %s\n" %mistake_index[0],
                             "When the datatype is integer, -1 is added : %s\n"%mistake_index[1],
                             "When the datatype is boolean, 0 is added : %s\n" % mistake_index[2]))
    f.close()

df1.save(os.path.join(path1,'%s_modified_final.xlsx'%(os.path.splitext(filename)[0])))
